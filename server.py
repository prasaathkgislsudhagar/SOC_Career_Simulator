#!/usr/bin/env python3
"""
server.py - Production SOC Career Simulator Backend with Excel Database (data/users.xlsx)
Features:
- Excel Database: data/users.xlsx with 'Users' worksheet
- Endpoint: POST /api/auth/register (Create Profile -> Save to Excel -> Auto Authenticate)
- Session Token Authorization (Bearer token)
- Scores & Progress Management (SQLite backend for performance, linked to Excel User ID)
- Static Asset Server with public file protection (blocks direct access to data/ and db/)
- No password system, no separate login endpoint
"""

import os
import re
import json
import sqlite3
import secrets
import threading
from datetime import datetime
from urllib.parse import parse_qs, unquote
from http.server import HTTPServer, SimpleHTTPRequestHandler
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DATA_DIR = os.path.join(BASE_DIR, 'data')
EXCEL_FILE = os.path.join(DATA_DIR, 'users.xlsx')
DB_DIR = os.path.join(BASE_DIR, 'db')
DB_FILE = os.path.join(DB_DIR, 'soc_platform.db')

EXCEL_LOCK = threading.Lock()

EXCEL_COLUMNS = [
    "User ID",
    "Username / Call-Sign",
    "Email",
    "Experience Level",
    "College / Organization",
    "Account Created At",
    "Last Login At",
    "Login Count",
    "Account Status"
]

def init_excel():
    """Initializes data/users.xlsx with the Users worksheet and styled headers if not present."""
    with EXCEL_LOCK:
        os.makedirs(DATA_DIR, exist_ok=True)
        if not os.path.exists(EXCEL_FILE):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Users"
            ws.append(EXCEL_COLUMNS)

            # Style Header Row
            header_fill = PatternFill(start_color="0B192C", end_color="0B192C", fill_type="solid")
            header_font = Font(name="Arial", size=11, bold=True, color="00F3FF")
            alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin', color='1E3E62'),
                right=Side(style='thin', color='1E3E62'),
                top=Side(style='thin', color='1E3E62'),
                bottom=Side(style='medium', color='00F3FF')
            )

            for col_num in range(1, len(EXCEL_COLUMNS) + 1):
                cell = ws.cell(row=1, column=col_num)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = alignment
                cell.border = thin_border
                # Set column width
                column_letter = openpyxl.utils.get_column_letter(col_num)
                ws.column_dimensions[column_letter].width = max(len(EXCEL_COLUMNS[col_num - 1]) + 6, 18)

            wb.save(EXCEL_FILE)
            print(f"[EXCEL] Created database: {EXCEL_FILE} with worksheet 'Users'")
        else:
            wb = openpyxl.load_workbook(EXCEL_FILE)
            if "Users" not in wb.sheetnames:
                ws = wb.create_sheet(title="Users")
                ws.append(EXCEL_COLUMNS)
                wb.save(EXCEL_FILE)
                print(f"[EXCEL] Added missing 'Users' worksheet to {EXCEL_FILE}")

def get_excel_users():
    """Reads all user rows from data/users.xlsx."""
    users = []
    if not os.path.exists(EXCEL_FILE):
        return users

    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)
    if "Users" not in wb.sheetnames:
        return users

    ws = wb["Users"]
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or not any(row):
            continue
        user_id = str(row[0] or '')
        username = str(row[1] or '')
        email = str(row[2] or '')
        experience = str(row[3] or '')
        college = str(row[4] or '')
        created_at = str(row[5] or '')
        last_login_at = str(row[6] or '')
        login_count = row[7] if row[7] is not None else 1
        account_status = str(row[8] or 'Active')

        users.append({
            'id': user_id,
            'username': username,
            'email': email,
            'experience': experience,
            'college': college,
            'organization': college,
            'createdAt': created_at,
            'lastLoginAt': last_login_at,
            'loginCount': login_count,
            'accountStatus': account_status
        })
    return users

def add_excel_user(user_data):
    """Appends a new user record to data/users.xlsx with proper formatting."""
    with EXCEL_LOCK:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb["Users"]
        
        row_values = [
            user_data['id'],
            user_data['username'],
            user_data['email'],
            user_data['experience'],
            user_data.get('college') or user_data.get('organization') or '',
            user_data['createdAt'],
            user_data['lastLoginAt'],
            user_data['loginCount'],
            user_data['accountStatus']
        ]
        ws.append(row_values)
        wb.save(EXCEL_FILE)
        print(f"[EXCEL] Registered new user '{user_data['username']}' ({user_data['id']}) in {EXCEL_FILE}")

def init_db():
    """Initializes SQLite database for session tokens, scores, and unlock progression."""
    os.makedirs(DB_DIR, exist_ok=True)
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()

    # Migrate tables if old integer schema exists
    cursor.execute("PRAGMA table_info(progress)")
    p_cols = cursor.fetchall()
    if p_cols:
        for col in p_cols:
            if col[1] == 'user_id' and 'INT' in str(col[2]).upper():
                cursor.execute("DROP TABLE IF EXISTS progress")
                cursor.execute("DROP TABLE IF EXISTS scores")
                cursor.execute("DROP TABLE IF EXISTS sessions")
                cursor.execute("DROP TABLE IF EXISTS users")
                break

    cursor.execute("PRAGMA table_info(sessions)")
    s_cols = [r[1] for r in cursor.fetchall()]
    if s_cols and 'username' not in s_cols:
        cursor.execute("DROP TABLE IF EXISTS sessions")

    # Sessions Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            user_id TEXT NOT NULL,
            username TEXT NOT NULL,
            email TEXT NOT NULL,
            experience TEXT,
            organization TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Scores Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS scores (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id TEXT NOT NULL,
            tier TEXT NOT NULL,
            scenario_id TEXT NOT NULL,
            score INTEGER NOT NULL,
            breakdown_json TEXT,
            created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    # Progress Table
    cursor.execute('''
        CREATE TABLE IF NOT EXISTS progress (
            user_id TEXT PRIMARY KEY,
            current_role TEXT DEFAULT 'tier1',
            unlocked_tiers_json TEXT,
            updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')

    conn.commit()
    conn.close()

def verify_token(token):
    """Verifies session token from SQLite sessions table and returns user profile."""
    if not token:
        return None
    conn = sqlite3.connect(DB_FILE)
    cursor = conn.cursor()
    cursor.execute('''
        SELECT user_id, username, email, experience, organization 
        FROM sessions 
        WHERE token = ?
    ''', (token,))
    row = cursor.fetchone()
    conn.close()
    if row:
        return {
            'id': row[0],
            'username': row[1],
            'email': row[2],
            'experience': row[3],
            'college': row[4] or '',
            'organization': row[4] or '',
            'accountStatus': 'Active'
        }
    return None

class SOCRequestHandler(SimpleHTTPRequestHandler):
    def translate_path(self, path):
        # Prevent path traversal outside project root
        return super().translate_path(path)

    def _send_json(self, data, status=200):
        self.send_response(status)
        self.send_header('Content-Type', 'application/json; charset=utf-8')
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.end_headers()
        self.wfile.write(json.dumps(data).encode('utf-8'))

    def do_OPTIONS(self):
        self.send_response(200)
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET, POST, OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type, Authorization')
        self.end_headers()

    def do_GET(self):
        parsed_path = unquote(self.path.split('?')[0])

        # Security check: Block direct public access to Excel database and SQLite db
        if parsed_path.startswith('/data') or parsed_path.startswith('/db') or 'users.xlsx' in parsed_path.lower():
            self.send_error(403, "Access Forbidden: Excel database and DB files cannot be accessed directly.")
            return

        if parsed_path == '/api/me':
            auth_header = self.headers.get('Authorization', '')
            token = auth_header.replace('Bearer ', '').strip()
            user = verify_token(token)
            if not user:
                return self._send_json({'error': 'Unauthorized'}, status=401)
            
            # Fetch User Scores & Progress from DB
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()

            cursor.execute("SELECT tier, scenario_id, score, breakdown_json FROM scores WHERE user_id = ?", (user['id'],))
            scores_rows = cursor.fetchall()
            
            tier_scores = {
                'tier1': {'scores': {}, 'average': 0},
                'tier2': {'scores': {}, 'average': 0},
                'tier3': {'scores': {}, 'average': 0},
                'manager': {'scores': {}, 'average': 0},
                'senior': {'scores': {}, 'average': 0}
            }
            for row in scores_rows:
                t, sc_id, sc_score, breakdown = row[0], row[1], row[2], row[3]
                if t in tier_scores:
                    tier_scores[t]['scores'][sc_id] = {'total': sc_score, 'breakdown': json.loads(breakdown) if breakdown else {}}

            for t in tier_scores:
                s_list = [s['total'] for s in tier_scores[t]['scores'].values()]
                tier_scores[t]['average'] = int(round(sum(s_list)/len(s_list))) if s_list else 0

            cursor.execute("SELECT current_role, unlocked_tiers_json FROM progress WHERE user_id = ?", (user['id'],))
            prog = cursor.fetchone()
            current_role = prog[0] if prog else 'tier1'
            unlocked_tiers = json.loads(prog[1]) if prog and prog[1] else {'tier1': True, 'tier2': False, 'tier3': False, 'manager': False, 'senior': False, 'master': False}

            conn.close()

            return self._send_json({
                'user': user,
                'currentRole': current_role,
                'tierScores': tier_scores,
                'unlockedTiers': unlocked_tiers
            })
        
        # Default static file handler
        return super().do_GET()

    def do_POST(self):
        parsed_path = unquote(self.path.split('?')[0])
        content_length = int(self.headers.get('Content-Length', 0))
        body = self.rfile.read(content_length).decode('utf-8') if content_length > 0 else '{}'
        
        try:
            data = json.loads(body) if body else {}
        except Exception:
            return self._send_json({'error': 'Malformed JSON payload.'}, status=400)

        # ----------------------------------------------------
        # CREATE PROFILE & REGISTER (Auto-Authenticate)
        # ----------------------------------------------------
        if parsed_path in ['/api/auth/register', '/api/register']:
            username = str(data.get('username') or '').strip()
            email = str(data.get('email') or '').strip()
            experience = str(data.get('experience') or 'Beginner (Student / Entry Analyst)').strip()
            college = str(data.get('college') or data.get('organization') or '').strip()

            # 1. Non-empty Validation
            if not username:
                return self._send_json({'error': 'Username / Call-Sign is required.'}, status=400)
            if not email:
                return self._send_json({'error': 'Corporate / Student Email is required.'}, status=400)
            if not experience:
                return self._send_json({'error': 'Cybersecurity Experience Level is required.'}, status=400)

            # 2. Email Format Validation
            email_regex = r'^[a-zA-Z0-9_.+-]+@[a-zA-Z0-9-]+\.[a-zA-Z0-9-.]+$'
            if not re.match(email_regex, email):
                return self._send_json({'error': 'Please enter a valid email address.'}, status=400)

            # 3. Read Existing Users from Excel & Duplicate Checks
            existing_users = get_excel_users()
            for u in existing_users:
                if u['username'].lower() == username.lower():
                    return self._send_json({
                        'error': 'Username already exists. Please choose another username.'
                    }, status=400)
                if u['email'].lower() == email.lower():
                    return self._send_json({
                        'error': 'An account with this email already exists.'
                    }, status=400)

            # 4. Generate User Details
            user_seq = len(existing_users) + 1
            user_id = f"USER-{1000 + user_seq}"
            now_str = datetime.now().strftime('%Y-%m-%d %H:%M:%S')

            new_user = {
                'id': user_id,
                'username': username,
                'email': email,
                'experience': experience,
                'college': college,
                'organization': college,
                'createdAt': now_str,
                'lastLoginAt': now_str,
                'loginCount': 1,
                'accountStatus': 'Active'
            }

            # 5. Append User to data/users.xlsx
            try:
                add_excel_user(new_user)
            except Exception as e:
                print(f"[EXCEL ERROR] Failed to write to {EXCEL_FILE}: {e}")
                return self._send_json({'error': 'Excel database write error. Please try again.'}, status=500)

            # 6. Automatic Authentication (Generate Session Token & Save in DB)
            token = secrets.token_hex(32)
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute(
                "INSERT INTO sessions (token, user_id, username, email, experience, organization) VALUES (?, ?, ?, ?, ?, ?)",
                (token, user_id, username, email, experience, college)
            )
            default_unlocks = json.dumps({'tier1': True, 'tier2': False, 'tier3': False, 'manager': False, 'senior': False, 'master': False})
            cursor.execute("INSERT OR REPLACE INTO progress (user_id, current_role, unlocked_tiers_json) VALUES (?, ?, ?)", (user_id, 'tier1', default_unlocks))
            conn.commit()
            conn.close()

            # 7. Return Token and User Profile
            return self._send_json({
                'token': token,
                'user': new_user
            }, status=201)

        # ----------------------------------------------------
        # RECORD SCENARIO SCORE & PROGRESSION
        # ----------------------------------------------------
        elif parsed_path == '/api/score':
            auth_header = self.headers.get('Authorization', '')
            token = auth_header.replace('Bearer ', '').strip()
            user = verify_token(token)
            if not user:
                return self._send_json({'error': 'Unauthorized'}, status=401)

            tier = data.get('tier')
            scenario_id = data.get('scenarioId')
            total_score = data.get('totalScore', 0)
            breakdown = data.get('breakdown', {})

            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            
            cursor.execute('''
                INSERT INTO scores (user_id, tier, scenario_id, score, breakdown_json)
                VALUES (?, ?, ?, ?, ?)
            ''', (user['id'], tier, scenario_id, total_score, json.dumps(breakdown)))

            # Evaluate Unlock Engine Server-Side
            cursor.execute("SELECT tier, scenario_id, score FROM scores WHERE user_id = ?", (user['id'],))
            all_scores = cursor.fetchall()
            
            tier_averages = {'tier1': 0, 'tier2': 0, 'tier3': 0, 'manager': 0, 'senior': 0}
            tier_counts = {'tier1': 0, 'tier2': 0, 'tier3': 0, 'manager': 0, 'senior': 0}
            
            for t_item in all_scores:
                t, sc_score = t_item[0], t_item[2]
                if t in tier_averages:
                    tier_averages[t] += sc_score
                    tier_counts[t] += 1
            
            for t in tier_averages:
                if tier_counts[t] > 0:
                    tier_averages[t] = int(round(tier_averages[t] / tier_counts[t]))

            cursor.execute("SELECT unlocked_tiers_json FROM progress WHERE user_id = ?", (user['id'],))
            prog_row = cursor.fetchone()
            unlocks = json.loads(prog_row[0]) if prog_row and prog_row[0] else {'tier1': True, 'tier2': False, 'tier3': False, 'manager': False, 'senior': False, 'master': False}

            newly_unlocked = []
            if tier_counts['tier1'] >= 2 and tier_averages['tier1'] >= 70 and not unlocks.get('tier2'):
                unlocks['tier2'] = True
                newly_unlocked.append('Tier 2 Analyst')
            if tier_counts['tier2'] >= 2 and tier_averages['tier2'] >= 75 and not unlocks.get('tier3'):
                unlocks['tier3'] = True
                newly_unlocked.append('Tier 3 Analyst')
            if tier_counts['tier3'] >= 2 and tier_averages['tier3'] >= 80 and not unlocks.get('manager'):
                unlocks['manager'] = True
                newly_unlocked.append('SOC Manager')
            if tier_counts['manager'] >= 2 and tier_averages['manager'] >= 85 and not unlocks.get('senior'):
                unlocks['senior'] = True
                newly_unlocked.append('Senior SOC Manager')
            if tier_counts['senior'] >= 2 and tier_averages['senior'] >= 90 and not unlocks.get('master'):
                unlocks['master'] = True
                newly_unlocked.append('SOC Master')

            cursor.execute("UPDATE progress SET unlocked_tiers_json = ? WHERE user_id = ?", (json.dumps(unlocks), user['id']))
            conn.commit()
            conn.close()

            return self._send_json({
                'success': True,
                'newlyUnlocked': newly_unlocked,
                'tierAverage': tier_averages.get(tier, 0)
            })

        # ----------------------------------------------------
        # LOGOUT
        # ----------------------------------------------------
        elif parsed_path == '/api/logout':
            auth_header = self.headers.get('Authorization', '')
            token = auth_header.replace('Bearer ', '').strip()
            conn = sqlite3.connect(DB_FILE)
            cursor = conn.cursor()
            cursor.execute("DELETE FROM sessions WHERE token = ?", (token,))
            conn.commit()
            conn.close()
            return self._send_json({'success': True})

        return self._send_json({'error': 'Endpoint not found.'}, status=404)

if __name__ == '__main__':
    init_excel()
    init_db()
    server_address = ('', 8000)
    httpd = HTTPServer(server_address, SOCRequestHandler)
    print("==========================================================")
    print("🛡️  SOC Career Simulator Backend Engine Active")
    print("🌐  URL:   http://localhost:8000")
    print("📊  EXCEL: data/users.xlsx (Worksheet: Users)")
    print("🔐  FLOW:  Create Profile -> Save to Excel -> Auto Login -> SOC")
    print("==========================================================")
    try:
        httpd.serve_forever()
    except KeyboardInterrupt:
        print("\nShutting down server.")
